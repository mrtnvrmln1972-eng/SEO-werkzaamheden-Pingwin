"""
generate_tracker.py

Genereert vindbaarheid-prioriteiten-tracker.xlsx in Pingwin huisstijl.
Leest dezelfde JSON-input als generate_report.js en schrijft een werkblad
"Bevindingen" plus een werkblad "Methodologie".

Aanroep:
    python3 generate_tracker.py <bevindingen.json> <output.xlsx>

Vereist: openpyxl. Installeer via:
    pip install openpyxl --break-system-packages
"""

from __future__ import annotations

import json
import sys
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# Pingwin design tokens (subset, alleen wat de tracker nodig heeft)
TOKENS = {
    "header_bg":       "FF222222",
    "header_text":     "FFFFFFFF",
    "row_white":       "FFFFFFFF",
    "row_grey":        "FFF7F7F7",
    "border_grey":     "FFEEEEEE",
    "accent":          "FFF15829",   # primair Pingwin accent
    "accent_secondary": "FFE7773F",  # CTA-oranje
    "accent_light":    "FFFFE5D6",
    "good_light":      "FFD6F0D6",
    "skip_grey":       "FFEEEEEE",
    "skip_text":       "FF888888",
    "body_text":       "FF181818",
    "heading_text":    "FF222222",
}

KOLOMMEN = [
    ("ID",                "bevinding_id",          6),
    ("Type",              "type",                  18),
    ("Titel",             "titel",                 35),
    ("URL",               "url",                   30),
    ("Zoekwoord",         "zoekwoord",             22),
    ("Volume",            "maandvolume",           9),
    ("Huidige positie",   "huidige_positie",       9),
    ("Target-positie",    "target_positie",        9),
    ("Impact",            "impact",                8),
    ("Effort",            "effort",                7),
    ("TTE",               "time_to_effect",        6),
    ("Confidence",        "confidence",            10),
    ("ROI-score",         "roi_score",             10),
    ("Tier",              "tier",                  14),
    ("Vervolg-skill",     "vervolg_skill",         25),
    ("Status",            "status",                12),
    ("Owner",             "owner",                 14),
    ("Datum",             "datum",                 12),
]

THIN_GREY = Side(style="thin", color=TOKENS["border_grey"])
GRID_BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def _font(bold: bool = False, color: str = TOKENS["body_text"], size: int = 10) -> Font:
    # Montserrat is op de meeste systemen niet als ttf geinstalleerd; openpyxl
    # zet de naam, en het systeem dat de xlsx opent valt terug op Calibri als
    # Montserrat niet bestaat. Dat is consistent met andere Pingwin-deliverables.
    return Font(name="Montserrat", bold=bold, color=color, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def _write_headers(ws, row: int = 1) -> None:
    for col_idx, (label, _key, _width) in enumerate(KOLOMMEN, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label.upper())
        cell.fill = _fill(TOKENS["header_bg"])
        cell.font = _font(bold=True, color=TOKENS["header_text"], size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER
    ws.row_dimensions[row].height = 28


def _set_column_widths(ws) -> None:
    for col_idx, (_label, _key, width) in enumerate(KOLOMMEN, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _row_styling(ws, row: int, tier: str, roi: float, status: str, is_top3: bool) -> None:
    # Alternating row fill
    base_fill = TOKENS["row_white"] if (row % 2 == 0) else TOKENS["row_grey"]

    for col_idx, (_label, key, _width) in enumerate(KOLOMMEN, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = _font(color=TOKENS["body_text"], size=10)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=key in ("titel", "zoekwoord", "vervolg_skill"),
        )
        cell.border = GRID_BORDER
        cell.fill = _fill(base_fill)

        # Tier-cel
        if key == "tier":
            if tier == "1":
                cell.fill = _fill(TOKENS["accent"])
                cell.font = _font(bold=True, color=TOKENS["row_white"], size=10)
            elif tier == "2":
                cell.fill = _fill(TOKENS["accent_secondary"])
                cell.font = _font(bold=True, color=TOKENS["row_white"], size=10)
            elif tier == "3":
                cell.fill = _fill(TOKENS["accent_light"])
            elif tier == "4":
                cell.fill = _fill(TOKENS["row_grey"])
            elif tier == "SKIP":
                cell.fill = _fill(TOKENS["skip_grey"])
                cell.font = _font(color=TOKENS["skip_text"], size=10)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=False, indent=0,
                )

        # ROI-cel
        if key == "roi_score":
            try:
                roi_val = float(roi)
            except (ValueError, TypeError):
                roi_val = 0.0
            if roi_val >= 5.0:
                cell.font = _font(bold=True, color=TOKENS["accent"], size=10)
            elif roi_val >= 2.0:
                cell.font = _font(bold=True, color=TOKENS["body_text"], size=10)
            elif roi_val < 0.5:
                cell.font = _font(color=TOKENS["skip_text"], size=10)

        # Status-cel
        if key == "status":
            if status == "in progress":
                cell.fill = _fill(TOKENS["accent_light"])
            elif status == "done":
                cell.fill = _fill(TOKENS["good_light"])
                cell.font = _font(color=TOKENS["body_text"], size=10)

    # Top-3 highlight: linkerrand 4pt accent op ID-cel
    if is_top3:
        id_cell = ws.cell(row=row, column=1)
        accent_side = Side(style="thick", color=TOKENS["accent"])
        id_cell.border = Border(
            left=accent_side, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY,
        )


def _value_for(b: dict, key: str) -> Any:
    val = b.get(key)
    if val is None:
        return ""
    return val


def write_bevindingen(ws, bevindingen: list[dict]) -> None:
    _write_headers(ws, row=1)
    _set_column_widths(ws)

    # Sorteer: tier oplopend (SKIP onderaan), dan ROI aflopend
    tier_order = {"1": 1, "2": 2, "3": 3, "4": 4, "SKIP": 99}

    def sort_key(b: dict) -> tuple:
        return (tier_order.get(b.get("tier", "SKIP"), 99), -float(b.get("roi_score", 0) or 0))

    sorted_b = sorted(bevindingen, key=sort_key)

    # Bepaal top-3 op ROI binnen niet-skip
    non_skip = [b for b in sorted_b if b.get("tier") != "SKIP"]
    top3_ids = {b.get("bevinding_id") for b in non_skip[:3]}

    today_str = date.today().strftime("%d-%m-%Y")
    for idx, b in enumerate(sorted_b, start=2):
        defaults = {
            "status": b.get("status", "open"),
            "owner":  b.get("owner", ""),
            "datum":  b.get("datum", today_str),
        }
        b_view = {**b, **defaults}

        for col_idx, (_label, key, _width) in enumerate(KOLOMMEN, start=1):
            ws.cell(row=idx, column=col_idx, value=_value_for(b_view, key))

        _row_styling(
            ws,
            row=idx,
            tier=str(b_view.get("tier", "")),
            roi=b_view.get("roi_score", 0),
            status=str(b_view.get("status", "")),
            is_top3=(b_view.get("bevinding_id") in top3_ids),
        )

    # Auto-filter en freeze panes
    last_col = get_column_letter(len(KOLOMMEN))
    ws.auto_filter.ref = f"A1:{last_col}{max(2, len(sorted_b) + 1)}"
    ws.freeze_panes = "A2"


def write_methodologie(ws, data: dict) -> None:
    rows = [
        ("Klant",            data.get("klant", "")),
        ("Datum scan",       data.get("datum", "")),
        ("Input-modus",      data.get("modus", "")),
        ("Propositie",       data.get("propositie", "")),
        ("CTR-curve-bron",   data.get("ctr_model", "Advanced Web Ranking 2024")),
    ]

    aggregaat = data.get("aggregaat", {})
    rows.extend([
        ("Items tier 1",         aggregaat.get("aantal_items_tier_1", 0)),
        ("Items tier 2",         aggregaat.get("aantal_items_tier_2", 0)),
        ("Items tier 3",         aggregaat.get("aantal_items_tier_3", 0)),
        ("Items tier 4",         aggregaat.get("aantal_items_tier_4", 0)),
        ("Items SKIP",           aggregaat.get("aantal_items_skip", 0)),
        ("Verwacht extra/maand", aggregaat.get("extra_bezoekers_per_maand", 0)),
    ])

    ai_prompts = data.get("ai_prompts", []) or []
    if ai_prompts:
        rows.append(("AI-prompts (lens 12)", "Onder deze rij opgesomd"))

    # Header rij
    ws.cell(row=1, column=1, value="KEY").fill = _fill(TOKENS["header_bg"])
    ws.cell(row=1, column=1).font = _font(bold=True, color=TOKENS["header_text"])
    ws.cell(row=1, column=2, value="VALUE").fill = _fill(TOKENS["header_bg"])
    ws.cell(row=1, column=2).font = _font(bold=True, color=TOKENS["header_text"])
    ws.row_dimensions[1].height = 24

    for r, (k, v) in enumerate(rows, start=2):
        c1 = ws.cell(row=r, column=1, value=k)
        c2 = ws.cell(row=r, column=2, value=v)
        for c in (c1, c2):
            c.font = _font(color=TOKENS["body_text"])
            c.fill = _fill(TOKENS["row_white"] if r % 2 == 0 else TOKENS["row_grey"])
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = GRID_BORDER

    # AI-prompts onderaan
    if ai_prompts:
        start = len(rows) + 2
        ws.cell(row=start, column=1, value="AI-prompts").font = _font(bold=True, color=TOKENS["heading_text"])
        for i, prompt in enumerate(ai_prompts, start=1):
            ws.cell(row=start + i, column=1, value=f"Prompt {i}")
            ws.cell(row=start + i, column=2, value=prompt)
            for col in (1, 2):
                cell = ws.cell(row=start + i, column=col)
                cell.font = _font(color=TOKENS["body_text"])
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = GRID_BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70


def build_workbook(data: dict) -> Workbook:
    wb = Workbook()

    ws_bev = wb.active
    ws_bev.title = "Bevindingen"

    # Stel alle bevindingen samen uit de tier-buckets en evt. losse "bevindingen"
    bevindingen: list[dict] = []
    for key in ("tier_1", "tier_2", "tier_3", "tier_4", "skip"):
        bevindingen.extend(data.get(key, []) or [])
    if not bevindingen and "bevindingen" in data:
        bevindingen = list(data["bevindingen"])

    write_bevindingen(ws_bev, bevindingen)

    ws_meth = wb.create_sheet("Methodologie")
    write_methodologie(ws_meth, data)

    return wb


def main() -> None:
    if len(sys.argv) != 3:
        print("Gebruik: python3 generate_tracker.py <bevindingen.json> <output.xlsx>")
        sys.exit(1)

    in_path, out_path = sys.argv[1], sys.argv[2]
    with open(in_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = build_workbook(data)
    wb.save(out_path)
    print(f"Tracker geschreven: {out_path}")


if __name__ == "__main__":
    main()
